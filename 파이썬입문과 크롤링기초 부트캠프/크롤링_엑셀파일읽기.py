import openpyxl
excel_file = openpyxl.load_workbook('test.xlsx') # 엑셀파일 오픈 ( load_workbook('파일명') )
print(excel_file.sheetnames) # 쉬트 이름 확인 (리스트 타입 return)

# 시트선택하기
excel_sheet = excel_file['상품정보']

# excel_sheet = excel_file.active # 시트가 한개면 .active 쌉가능

# 시트 안에 데이터 읽기
for row in excel_sheet.rows: # 각 행을 가지고옴 .rows
    print(row[0].value, row[1].value)

# 오픈한 엑셀 파일 닫기
excel_file.close()


